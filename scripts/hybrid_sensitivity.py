"""
PGAI Group 5 — Hybrid Weight Sensitivity Analysis
===================================================
Tests 3 hybrid weight configurations on 12 failing questions:
  - 0.0 = pure semantic search
  - 0.5 = balanced (current baseline)
  - 1.0 = pure keyword search

Also tests chunk size impact on T2-03 specifically.

Run from rag-skeleton directory:
  python hybrid_sensitivity.py
"""

import re
import json
import requests
import time
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ============================================================
# CONFIGURATION
# ============================================================
TOGETHER_API_KEY  = ""
OPENAI_API_KEY    = ""
RAG_API_URL       = "http://localhost:7860"
QUESTIONS_PATH    = r"C:\Users\DELL\OneDrive - Indian Institute of Management\Desktop\Term 4\P&GAI\End Term Project\Questions_PGAI_Group_5_FINAL.xlsx"
INPUT_SHEET       = "All Questions"
OUTPUT_EXCEL      = "hybrid_sensitivity_results.xlsx"
CHECKPOINT_CSV    = "hybrid_sensitivity_checkpoint.csv"

EMBEDDING_MODEL   = "MiniLM-L6 (fast, 384d)"
ANSWER_MODEL      = "gpt-oss-20b (Together AI)"
TOP_K             = 6      # Increased from 4 to capture more chunks
DELAY             = 1.5

# 12 failing question IDs (score < 6 in Phase 2 evaluation)
FAILING_IDS = [
    "BANKING-T2-03", "BANKING-T2-10",
    "BANKING-T3-04",
    "BANKING-T4-02", "BANKING-T4-04", "BANKING-T4-05",
    "BANKING-T4-06", "BANKING-T4-09",
    "BANKING-T5-01", "BANKING-T5-02", "BANKING-T5-06", "BANKING-T5-10",
]

# 3 hybrid weight configurations to test
HYBRID_CONFIGS = [
    {"hybrid_weight": 0.0, "label": "Pure Semantic (0.0)"},
    {"hybrid_weight": 0.5, "label": "Balanced (0.5) — current"},
    {"hybrid_weight": 1.0, "label": "Pure Keyword (1.0)"},
]

# Phase 2 baseline scores for comparison
BASELINE_SCORES = {
    "BANKING-T2-03": 5.17, "BANKING-T2-10": 4.83,
    "BANKING-T3-04": 5.50,
    "BANKING-T4-02": 1.83, "BANKING-T4-04": 5.50,
    "BANKING-T4-05": 5.33, "BANKING-T4-06": 5.67,
    "BANKING-T4-09": 5.50,
    "BANKING-T5-01": 5.17, "BANKING-T5-02": 5.67,
    "BANKING-T5-06": 5.50, "BANKING-T5-10": 5.17,
}

# ============================================================
# ENCODING FIX (same as run_rag_benchmark.py)
# ============================================================
def fix_response(val):
    if not isinstance(val, str):
        return val
    try:
        val = val.encode('latin-1').decode('utf-8')
    except:
        pass
    val = val.replace('\u202f', ' ')
    val = val.replace('\xa0', ' ')
    val = val.replace('\xad', '-')
    val = val.replace('\u2011', '-')
    val = val.replace('【', '[').replace('】', ']')
    val = re.sub(r'\*\*(.*?)\*\*', r'\1', val)
    val = re.sub(r'\*(.*?)\*', r'\1', val)
    return val

# ============================================================
# EVALUATOR CLIENTS (same as evaluate_rag.py)
# ============================================================
EVALUATORS_CFG = [
    {
        "name": "gpt-5-nano",
        "model": "gpt-5-nano",
        "base_url": None,
        "api_key": OPENAI_API_KEY,
        "reasoning": True,
    },
    {
        "name": "Gemma-3N-E4B",
        "model": "google/gemma-3n-E4B-it",
        "base_url": "https://api.together.xyz/v1",
        "api_key": TOGETHER_API_KEY,
        "reasoning": False,
    },
]

EVAL_CLIENTS = {}
for cfg in EVALUATORS_CFG:
    if cfg["base_url"]:
        EVAL_CLIENTS[cfg["name"]] = OpenAI(
            api_key=cfg["api_key"], base_url=cfg["base_url"])
    else:
        EVAL_CLIENTS[cfg["name"]] = OpenAI(api_key=cfg["api_key"])

EVAL_PROFILES = [
    {
        "id": "E1", "label": "Strict RBI Examiner",
        "system": """You are a strict RBI regulatory examiner scoring AI responses.
Rubric (max 8 pts):
Factual Accuracy (0-4): 4=No errors|3=Minor omissions|2=Material error|1=Hallucinated|0=No answer
Completeness (0-2): 2=All aspects|1=Core only|0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident|1=Over/under|0=Confident hallucination
Respond ONLY in JSON: {"factual_accuracy":<0-4>,"completeness":<0-2>,"confidence_calibration":<0-2>,"total":<sum>,"reasoning":"<one sentence>"}"""
    },
    {
        "id": "E2", "label": "Balanced Compliance Analyst",
        "system": """You are a balanced compliance analyst scoring AI responses.
Rubric (max 8 pts):
Factual Accuracy (0-4): 4=No errors|3=Minor omissions|2=Material error|1=Hallucinated|0=No answer
Completeness (0-2): 2=All aspects|1=Core only|0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident|1=Over/under|0=Confident hallucination
Respond ONLY in JSON: {"factual_accuracy":<0-4>,"completeness":<0-2>,"confidence_calibration":<0-2>,"total":<sum>,"reasoning":"<one sentence>"}"""
    },
    {
        "id": "E3", "label": "Practical Compliance Officer",
        "system": """You are a senior compliance officer scoring AI responses for real-world usefulness.
Rubric (max 8 pts):
Factual Accuracy (0-4): 4=No errors|3=Minor omissions|2=Material error|1=Hallucinated|0=No answer
Completeness (0-2): 2=All aspects|1=Core only|0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident|1=Over/under|0=Confident hallucination
Respond ONLY in JSON: {"factual_accuracy":<0-4>,"completeness":<0-2>,"confidence_calibration":<0-2>,"total":<sum>,"reasoning":"<one sentence>"}"""
    },
]

def score_response(question, expected, response):
    """Score one response using same cross-model method as Phase 2."""
    profile_scores = {}
    for ev in EVAL_PROFILES:
        ext_scores = []
        for cfg in EVALUATORS_CFG:
            client = EVAL_CLIENTS[cfg["name"]]
            user_msg = f"""QUESTION:\n{question}\n\nEXPECTED ANSWER:\n{expected}\n\nRAG RESPONSE:\n{response}\n\nScore using the rubric."""
            params = {
                "model": cfg["model"],
                "messages": [
                    {"role": "system", "content": ev["system"]},
                    {"role": "user", "content": user_msg}
                ],
            }
            if cfg["reasoning"]:
                params["max_completion_tokens"] = 5000
            else:
                params["max_tokens"] = 512
                params["temperature"] = 0
            try:
                resp  = client.chat.completions.create(**params)
                raw   = resp.choices[0].message.content or ""
                match = re.search(r'\{.*?\}', raw, re.DOTALL)
                if match:
                    result = json.loads(match.group())
                    total  = (max(0,min(4,int(result.get("factual_accuracy",0)))) +
                              max(0,min(2,int(result.get("completeness",0)))) +
                              max(0,min(2,int(result.get("confidence_calibration",0)))))
                    ext_scores.append(total)
            except Exception as e:
                print(f"         scorer error: {str(e)[:60]}")
            time.sleep(DELAY)

        if ext_scores:
            profile_scores[ev["id"]] = round(sum(ext_scores)/len(ext_scores), 2)

    if profile_scores:
        return round(sum(profile_scores.values())/len(profile_scores), 2), profile_scores
    return None, {}

# ============================================================
# STEP 1: Verify server
# ============================================================
print("=" * 65)
print("  PGAI GROUP 5 — HYBRID SENSITIVITY ANALYSIS")
print("=" * 65)
print(f"\n  Testing {len(FAILING_IDS)} failing questions × {len(HYBRID_CONFIGS)} configs")
print(f"  Total RAG calls    : {len(FAILING_IDS) * len(HYBRID_CONFIGS)}")
print(f"  Total scoring calls: {len(FAILING_IDS) * len(HYBRID_CONFIGS) * len(EVAL_PROFILES) * 2}")
print()

print("🔍 Checking RAG server...")
try:
    r = requests.get(f"{RAG_API_URL}/answer-models", timeout=5)
    print(f"✅ Server ready\n")
except Exception as e:
    print(f"❌ Server not reachable — start uvicorn first\n   {e}")
    exit(1)

# ============================================================
# STEP 2: Load questions
# ============================================================
print(f"📂 Loading questions...")
df_all = pd.read_excel(QUESTIONS_PATH, sheet_name=INPUT_SHEET)
df_all.columns = [c.strip() for c in df_all.columns]
df_fail = df_all[df_all["question_id"].isin(FAILING_IDS)].reset_index(drop=True)
print(f"✅ {len(df_fail)} failing questions loaded\n")

# ============================================================
# STEP 3: Resume from checkpoint
# ============================================================
done_keys = set()
all_results = []
if os.path.exists(CHECKPOINT_CSV):
    ck = pd.read_csv(CHECKPOINT_CSV, encoding="utf-8-sig")
    done_keys = set(zip(ck["question_id"], ck["hybrid_weight"].astype(str)))
    all_results = ck.to_dict("records")
    print(f"♻️  Resuming — {len(done_keys)} already done\n")
else:
    print("🚀 Fresh start\n")

# ============================================================
# STEP 4: Run all combinations
# ============================================================
start_time = time.time()
total_runs = len(df_fail) * len(HYBRID_CONFIGS)
run_num    = 0

for _, row in df_fail.iterrows():
    q_id     = row["question_id"]
    tier     = row["tier"]
    q_text   = str(row["question_text"]).strip()
    expected = str(row.get("expected_answer","")).strip()
    baseline = BASELINE_SCORES.get(q_id, "")

    for cfg in HYBRID_CONFIGS:
        hw    = cfg["hybrid_weight"]
        label = cfg["label"]
        key   = (q_id, str(hw))
        run_num += 1

        if key in done_keys:
            print(f"[{run_num}/{total_runs}] {q_id} | hw={hw} ⏭ skipped")
            continue

        print(f"\n[{run_num}/{total_runs}] {q_id} | Tier {tier} | {label}")

        # ── RAG call ──
        payload = {
            "question":        q_text,
            "embedding_model": EMBEDDING_MODEL,
            "answer_model":    ANSWER_MODEL,
            "top_k":           TOP_K,
            "hybrid_weight":   hw,
        }
        try:
            resp     = requests.post(f"{RAG_API_URL}/query", json=payload, timeout=60)
            data     = resp.json()
            rag_ans  = fix_response(data.get("answer",""))
            chunks   = data.get("chunks",[])
            top_src  = "; ".join([f"{c.get('filename','?')} ({c.get('score',0):.3f})"
                                   for c in chunks[:3]])
            top_score= chunks[0].get("score",0) if chunks else 0
            rag_status = "OK" if rag_ans else "EMPTY"
        except Exception as e:
            rag_ans    = ""
            top_src    = ""
            top_score  = 0
            rag_status = f"ERROR: {str(e)[:60]}"

        print(f"   RAG: {rag_status} | top_score={top_score:.3f}")
        if rag_ans:
            print(f"   Answer: {rag_ans[:100]}...")

        # ── Score ──
        if rag_ans:
            print(f"   Scoring...")
            final_score, profile_scores = score_response(q_text, expected, rag_ans)
        else:
            final_score    = None
            profile_scores = {}

        improvement = round(final_score - baseline, 2) if final_score and baseline else ""

        all_results.append({
            "question_id":   q_id,
            "tier":          tier,
            "question_text": q_text,
            "hybrid_weight": hw,
            "config_label":  label,
            "top_k":         TOP_K,
            "rag_response":  rag_ans,
            "top_sources":   top_src,
            "top_similarity":round(top_score,3),
            "E1":            profile_scores.get("E1",""),
            "E2":            profile_scores.get("E2",""),
            "E3":            profile_scores.get("E3",""),
            "final_score":   final_score if final_score else "",
            "baseline_score":baseline,
            "improvement":   improvement,
            "rag_status":    rag_status,
        })

        # Save checkpoint after every run
        pd.DataFrame(all_results).to_csv(CHECKPOINT_CSV, index=False, encoding="utf-8-sig")

        elapsed   = time.time() - start_time
        remaining = (elapsed / run_num) * (total_runs - run_num) if run_num else 0
        print(f"   Score: {final_score}/8 (baseline: {baseline}, Δ {improvement})")
        print(f"   💾 | ⏱ {elapsed/60:.1f}m | ~{remaining/60:.1f}m remaining")

results_df = pd.DataFrame(all_results)

# ============================================================
# STEP 5: Print summary
# ============================================================
print(f"\n{'='*65}")
print("  HYBRID SENSITIVITY — SUMMARY")
print(f"{'='*65}")
for cfg in HYBRID_CONFIGS:
    hw    = cfg["hybrid_weight"]
    label = cfg["label"]
    sub   = results_df[results_df["hybrid_weight"]==hw]
    avg   = pd.to_numeric(sub["final_score"],errors="coerce").mean()
    base  = pd.to_numeric(sub["baseline_score"],errors="coerce").mean()
    print(f"\n  {label}")
    print(f"    Avg score    : {avg:.2f}/8")
    print(f"    Avg baseline : {base:.2f}/8")
    print(f"    Avg Δ        : {avg-base:+.2f}")
print(f"\n{'='*65}\n")

# ============================================================
# STEP 6: Write Excel
# ============================================================
print(f"💾 Writing: {OUTPUT_EXCEL}")
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Sensitivity Results"

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
WHITE_BOLD= Font(name="Arial", bold=True, color="FFFFFF", size=11)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP      = Alignment(wrap_text=True, vertical="top")

TIER_CLR  = {"1":"EBF5FB","2":"E9F7EF","3":"FEF9E7","4":"FEF5E7","5":"FDEDEC"}
HW_CLR    = {"0.0":"D5F5E3","0.5":"D6EAF8","1.0":"FDEBD0"}
HW_HDR    = {"0.0":"1E8449","0.5":"1A5276","1.0":"6E2F00"}
def mf(h): return PatternFill("solid", fgColor=h)

HEADERS = [
    ("question_id",   "1F4E79"),("tier",         "1F4E79"),
    ("question_text", "1F4E79"),("config_label",  "1F4E79"),
    ("hybrid_weight", "1F4E79"),("top_k",         "1F4E79"),
    ("rag_response",  "1F4E79"),("top_sources",   "1F4E79"),
    ("top_similarity","1F4E79"),
    ("E1","1F4E79"),("E2","1F4E79"),("E3","1F4E79"),
    ("final_score",   "1F4E79"),("baseline_score","1F4E79"),
    ("improvement",   "1F4E79"),
]
for c,(label,color) in enumerate(HEADERS,start=1):
    cell=ws.cell(row=1,column=c,value=label)
    cell.font=HDR_FONT; cell.fill=mf(color); cell.alignment=CENTER
ws.row_dimensions[1].height=30

for ri,row in results_df.iterrows():
    er    = ri+2
    hw    = str(row.get("hybrid_weight",""))
    t_str = str(row.get("tier","")).strip()
    fill  = mf(HW_CLR.get(hw, TIER_CLR.get(t_str,"FFFFFF")))
    data_row = [
        row.get("question_id",""),  row.get("tier",""),
        row.get("question_text",""),row.get("config_label",""),
        row.get("hybrid_weight",""),row.get("top_k",""),
        row.get("rag_response",""), row.get("top_sources",""),
        row.get("top_similarity",""),
        row.get("E1",""),row.get("E2",""),row.get("E3",""),
        row.get("final_score",""),  row.get("baseline_score",""),
        row.get("improvement",""),
    ]
    for ci,val in enumerate(data_row,start=1):
        val  = "" if pd.isna(val) else val
        cell = ws.cell(row=er,column=ci,value=val)
        is_score = ci in (10,11,12,13,14,15)
        cell.font      = WHITE_BOLD if ci==13 else BOLD_FONT if is_score else BODY_FONT
        cell.fill      = mf(HW_HDR.get(hw,"1F4E79")) if ci==13 else fill
        cell.alignment = CENTER if ci not in (3,7,8) else WRAP

WIDTHS={1:16,2:6,3:42,4:22,5:13,6:7,7:55,8:40,9:13,
        10:8,11:8,12:8,13:12,14:13,15:12}
for c,w in WIDTHS.items():
    ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes="A2"

# ── Summary tab ──
ws2 = wb.create_sheet("Summary")
sum_hdrs = ["Config","Hybrid Weight","Avg Score /8","Avg Baseline","Avg Δ","Best for Tier"]
for c,h in enumerate(sum_hdrs,start=1):
    cell=ws2.cell(row=1,column=c,value=h)
    cell.font=HDR_FONT; cell.fill=mf("1F4E79"); cell.alignment=CENTER
ws2.row_dimensions[1].height=28

for ri,cfg in enumerate(HYBRID_CONFIGS,start=2):
    hw   = cfg["hybrid_weight"]
    sub  = results_df[results_df["hybrid_weight"]==hw]
    avg  = pd.to_numeric(sub["final_score"],errors="coerce").mean()
    base = pd.to_numeric(sub["baseline_score"],errors="coerce").mean()
    # Best tier for this config
    best_tier = ""
    best_imp  = -99
    for tier in [2,3,4,5]:
        t_sub = sub[sub["tier"]==tier]
        t_imp = pd.to_numeric(t_sub["improvement"],errors="coerce").mean()
        if not pd.isna(t_imp) and t_imp > best_imp:
            best_imp  = t_imp
            best_tier = f"T{tier} (+{t_imp:.2f})"
    vals=[cfg["label"],hw,round(avg,2),round(base,2),round(avg-base,2),best_tier]
    fill=mf(HW_CLR.get(str(hw),"FFFFFF"))
    for c,v in enumerate(vals,start=1):
        cell=ws2.cell(row=ri,column=c,value=v)
        cell.font=BOLD_FONT if c in (3,5) else BODY_FONT
        cell.fill=fill; cell.alignment=CENTER
for c in range(1,len(sum_hdrs)+1):
    ws2.column_dimensions[get_column_letter(c)].width=20
ws2.freeze_panes="A2"

wb.save(OUTPUT_EXCEL)
print(f"✅ Saved: {OUTPUT_EXCEL}")
print(f"✅ Checkpoint: {CHECKPOINT_CSV}")
print(f"\n🎉 Hybrid sensitivity analysis complete!")
print(f"   Open {OUTPUT_EXCEL} → Summary tab for comparison")
