"""
PGAI Group 5 - Level 3: Query Decomposition for T4
Decomposes multi-step questions into sub-questions,
retrieves chunks for each, combines and generates answer.
"""

import re, json, requests, time, os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ── CONFIG ──
TOGETHER_API_KEY = ""
OPENAI_API_KEY   = ""
RAG_API_URL      = "http://localhost:7860"
QUESTIONS_PATH   = r"C:\Users\DELL\OneDrive - Indian Institute of Management\Desktop\Term 4\P&GAI\End Term Project\Questions_PGAI_Group_5_FINAL.xlsx"
INPUT_SHEET      = "All Questions"
OUTPUT_EXCEL     = "query_decomposition_results.xlsx"
CHECKPOINT_CSV   = "query_decomposition_checkpoint.csv"

EMBEDDING_MODEL  = "MiniLM-L6 (fast, 384d)"
ANSWER_MODEL     = "gpt-oss-20b (Together AI)"
HYBRID_WEIGHT    = 0.0
TOP_K            = 6
DELAY            = 1.5

TARGET_IDS = [
    "BANKING-T4-02", "BANKING-T4-04", "BANKING-T4-05",
    "BANKING-T4-06", "BANKING-T4-09",
]

BASELINE_SCORES = {
    "BANKING-T4-02": 1.83, "BANKING-T4-04": 5.50,
    "BANKING-T4-05": 5.33, "BANKING-T4-06": 5.67,
    "BANKING-T4-09": 5.50,
}

# Pre-defined sub-questions (avoids API truncation issues)
SUBQUESTIONS = {
    "BANKING-T4-02": [
        "What RBI rules govern LSP role in digital loan disbursement and repayment?",
        "What are the disclosure requirements for EIR calculation including upfront fees?",
        "What are the KFS delivery timelines and borrower consent rules in digital lending?"
    ],
    "BANKING-T4-04": [
        "What NBFC concentration limits apply to infrastructure project loans under RBI?",
        "What steps must an NBFC take when a loan exceeds DCCO due to court proceedings?",
        "What classification rules apply when NBFC breaches concentration norms?"
    ],
    "BANKING-T4-05": [
        "How does RBI define NPA classification for funded facilities with same borrower?",
        "When does an unpaid derivative receivable trigger NPA for other facilities?",
        "What provisioning requirements apply when NBFC classifies accounts as NPA?"
    ],
    "BANKING-T4-06": [
        "What RBI timeline applies for NBFC identification of SMA and NPA accounts?",
        "What corrective actions must NBFC take after discovering delayed NPA identification?",
        "What reporting obligations apply to NBFC under PCA framework for asset quality?"
    ],
    "BANKING-T4-09": [
        "What are the RBI rules for PCFC repayment timelines for exporters?",
        "How must a bank handle PCFC when export proceeds are delayed beyond due date?",
        "What NPA classification rules apply to overdue PCFC accounts under RBI guidelines?"
    ],
}

# ── CLIENTS ──
together_client = OpenAI(api_key=TOGETHER_API_KEY, base_url="https://api.together.xyz/v1")
nano_client     = OpenAI(api_key=OPENAI_API_KEY)
gemma_client    = OpenAI(api_key=TOGETHER_API_KEY, base_url="https://api.together.xyz/v1")

# ── ENCODING FIX ──
def fix_text(val):
    if not isinstance(val, str):
        return val
    try:
        val = val.encode('latin-1').decode('utf-8')
    except:
        pass
    val = val.replace('\u202f',' ').replace('\xa0',' ')
    val = val.replace('\xad','-').replace('\u2011','-')
    val = val.replace('\u300c','[').replace('\u300d',']')
    val = re.sub(r'\*\*(.*?)\*\*', r'\1', val)
    return val

# ── RETRIEVE FOR ONE SUB-QUESTION ──
def retrieve_chunks(sub_question):
    payload = {
        "question": sub_question,
        "embedding_model": EMBEDDING_MODEL,
        "answer_model": ANSWER_MODEL,
        "top_k": TOP_K,
        "hybrid_weight": HYBRID_WEIGHT,
    }
    try:
        resp   = requests.post(f"{RAG_API_URL}/query", json=payload, timeout=60)
        data   = resp.json()
        return data.get("chunks", [])
    except Exception as e:
        print(f"      Retrieval error: {e}")
        return []

# ── GENERATE FROM COMBINED CHUNKS ──
def generate_answer(question, all_chunks):
    seen, unique = set(), []
    for c in all_chunks:
        txt = c.get("text", c.get("content",""))
        if txt and txt not in seen:
            seen.add(txt); unique.append(c)
    if not unique:
        return "", []
    context = "\n\n".join(f"[{i+1}] {c.get('text',c.get('content',''))}" for i,c in enumerate(unique))
    prompt  = ("Answer the question using only the context below. "
               "If the context does not contain the answer, say so explicitly. "
               "Cite sources inline like [1], [2].\n\n"
               f"Context:\n{context}\n\nQuestion: {question}")
    try:
        resp = together_client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[{"role":"user","content":prompt}],
            max_completion_tokens=2000,
        )
        return fix_text(resp.choices[0].message.content or ""), unique
    except Exception as e:
        print(f"      Generation error: {e}")
        return "", unique

# ── SCORE ONE RESPONSE ──
RUBRIC = ("Rubric (max 8 pts): "
          "Factual Accuracy (0-4): 4=No errors, 3=Minor omissions, 2=Material error, 1=Hallucinated, 0=No answer. "
          "Completeness (0-2): 2=All aspects, 1=Core only, 0=Superficial. "
          "Calibration (0-2): 2=Appropriately confident, 1=Over/under confident, 0=Confident hallucination. "
          'Respond ONLY with JSON like: {"factual_accuracy":3,"completeness":2,"confidence_calibration":2,"total":7,"reasoning":"one sentence"}')

SCORERS = [
    {"name":"gpt-5-nano",   "model":"gpt-5-nano",            "client":None, "reasoning":True},
    {"name":"Gemma-3N-E4B", "model":"google/gemma-3n-E4B-it","client":None, "reasoning":False},
]
SCORERS[0]["client"] = nano_client
SCORERS[1]["client"] = gemma_client

PROFILES = [
    ("E1", "Strict RBI Examiner — zero tolerance for deviation from exact RBI text. " + RUBRIC),
    ("E2", "Balanced Compliance Analyst — fair partial credit for directionally correct answers. " + RUBRIC),
    ("E3", "Practical Compliance Officer — would this answer lead to the right regulatory decision? " + RUBRIC),
]

def score_response(question, expected, response):
    profile_scores = {}
    for eid, system in PROFILES:
        ext = []
        for scorer in SCORERS:
            params = {
                "model":    scorer["model"],
                "messages": [
                    {"role":"system","content":system},
                    {"role":"user","content":f"QUESTION:\n{question}\n\nEXPECTED:\n{expected}\n\nRESPONSE:\n{response}\n\nScore this."}
                ],
            }
            if scorer["reasoning"]:
                params["max_completion_tokens"] = 5000
            else:
                params["max_tokens"] = 512
                params["temperature"] = 0
            try:
                resp  = scorer["client"].chat.completions.create(**params)
                raw   = resp.choices[0].message.content or ""
                match = re.search(r'\{[^{}]+\}', raw, re.DOTALL)
                if match:
                    r     = json.loads(match.group())
                    total = (max(0,min(4,int(r.get("factual_accuracy",0)))) +
                             max(0,min(2,int(r.get("completeness",0)))) +
                             max(0,min(2,int(r.get("confidence_calibration",0)))))
                    ext.append(total)
            except Exception as e:
                print(f"         scorer error: {str(e)[:60]}")
            time.sleep(DELAY)
        if ext:
            profile_scores[eid] = round(sum(ext)/len(ext), 2)
    if profile_scores:
        return round(sum(profile_scores.values())/len(profile_scores), 2), profile_scores
    return None, {}

# ── MAIN ──
print("=" * 65)
print("  PGAI GROUP 5 - LEVEL 3: QUERY DECOMPOSITION")
print("=" * 65)
print(f"\n  5 T4 questions | hw={HYBRID_WEIGHT} | Top-K={TOP_K}")
print(f"  Sub-questions pre-defined (3 per question)\n")

print("Checking RAG server...")
try:
    requests.get(f"{RAG_API_URL}/answer-models", timeout=5)
    print("Server ready\n")
except Exception as e:
    print(f"Server not reachable: {e}"); exit(1)

print("Loading questions...")
df_all    = pd.read_excel(QUESTIONS_PATH, sheet_name=INPUT_SHEET)
df_all.columns = [c.strip() for c in df_all.columns]
df_target = df_all[df_all["question_id"].isin(TARGET_IDS)].reset_index(drop=True)
print(f"{len(df_target)} T4 questions loaded\n")

done_ids, all_results = set(), []
if os.path.exists(CHECKPOINT_CSV):
    ck = pd.read_csv(CHECKPOINT_CSV, encoding="utf-8-sig")
    done_ids    = set(ck["question_id"].tolist())
    all_results = ck.to_dict("records")
    print(f"Resuming - {len(done_ids)} already done\n")
else:
    print("Fresh start\n")

start_time = time.time()

for i, row in df_target.iterrows():
    q_id     = row["question_id"]
    tier     = row["tier"]
    q_text   = str(row["question_text"]).strip()
    expected = str(row.get("expected_answer","")).strip()
    baseline = BASELINE_SCORES.get(q_id,"")

    if q_id in done_ids:
        print(f"[{i+1}/{len(df_target)}] {q_id} skipped")
        continue

    print(f"\n[{i+1}/{len(df_target)}] {q_id} | Tier {tier}")

    # Step 1: Get sub-questions
    subs = SUBQUESTIONS.get(q_id, [])
    print(f"  Step 1 - Sub-questions:")
    for j,s in enumerate(subs):
        print(f"    {j+1}. {s}")

    # Step 2: Retrieve chunks for each sub-question
    print(f"\n  Step 2 - Retrieving chunks...")
    all_chunks = []
    for j, sq in enumerate(subs):
        chunks = retrieve_chunks(sq)
        print(f"    Sub-Q{j+1}: {len(chunks)} chunks | top={chunks[0].get('filename','?') if chunks else 'none'}")
        all_chunks.extend(chunks)
        time.sleep(1)

    # Deduplicate
    seen, unique = set(), []
    for c in all_chunks:
        txt = c.get("text", c.get("content",""))
        if txt and txt not in seen:
            seen.add(txt); unique.append(c)
    print(f"    Total unique chunks: {len(unique)}")

    # Step 3: Generate
    print(f"\n  Step 3 - Generating answer...")
    answer, used = generate_answer(q_text, unique)
    if not answer:
        print(f"  Generation failed — skipping")
        continue
    print(f"  Answer ({len(answer)} chars): {answer[:120]}...")

    # Step 4: Score
    print(f"\n  Step 4 - Scoring...")
    final_score, ps = score_response(q_text, expected, answer)
    improvement = round(final_score - float(baseline), 2) if final_score and baseline else ""
    print(f"  Result: {final_score}/8 (baseline={baseline}, delta={improvement})")

    # Sources
    sources = "; ".join(set(c.get("filename","?") for c in unique[:6]))

    all_results.append({
        "question_id":     q_id,
        "tier":            tier,
        "question_text":   q_text,
        "expected_answer": expected,
        "sub_q1":          subs[0] if len(subs)>0 else "",
        "sub_q2":          subs[1] if len(subs)>1 else "",
        "sub_q3":          subs[2] if len(subs)>2 else "",
        "unique_chunks":   len(unique),
        "sources":         sources,
        "final_answer":    answer,
        "E1":              ps.get("E1",""),
        "E2":              ps.get("E2",""),
        "E3":              ps.get("E3",""),
        "final_score":     final_score or "",
        "baseline_score":  baseline,
        "improvement":     improvement,
    })

    pd.DataFrame(all_results).to_csv(CHECKPOINT_CSV, index=False, encoding="utf-8-sig")
    elapsed = time.time() - start_time
    print(f"  Saved | {elapsed/60:.1f}m elapsed")

results_df = pd.DataFrame(all_results)

print(f"\n{'='*65}")
print("  RESULTS")
print(f"{'='*65}")
for _,r in results_df.iterrows():
    imp = r.get('improvement','')
    print(f"  {r['question_id']}: {r['baseline_score']} -> {r['final_score']} ({imp:+.2f})" if isinstance(imp,float) else f"  {r['question_id']}: {r.get('final_score','')}")

# Write Excel
print(f"\nWriting: {OUTPUT_EXCEL}")
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Query Decomposition"

HF  = Font(name="Arial",bold=True,color="FFFFFF",size=10)
BF  = Font(name="Arial",size=10)
WB  = Font(name="Arial",bold=True,color="FFFFFF",size=11)
C   = Alignment(horizontal="center",vertical="center",wrap_text=True)
W   = Alignment(wrap_text=True,vertical="top")
def mf(h): return PatternFill("solid",fgColor=h)

HDRS = [
    ("question_id","1F4E79"),("tier","1F4E79"),
    ("question_text","1F4E79"),("expected_answer","1F4E79"),
    ("sub_q1","784212"),("sub_q2","784212"),("sub_q3","784212"),
    ("unique_chunks","784212"),("sources","784212"),
    ("final_answer","1A5276"),
    ("E1","1A5276"),("E2","1A5276"),("E3","1A5276"),
    ("final_score","2E86C1"),("baseline_score","1F4E79"),("improvement","1F4E79"),
]
for c,(h,col) in enumerate(HDRS,start=1):
    cell=ws.cell(row=1,column=c,value=h)
    cell.font=HF; cell.fill=mf(col); cell.alignment=C
ws.row_dimensions[1].height=30

for ri,row in results_df.iterrows():
    er=ri+2
    vals=[row.get(h[0],"") for h in HDRS]
    for ci,val in enumerate(vals,start=1):
        val="" if pd.isna(val) else val
        cell=ws.cell(row=er,column=ci,value=val)
        cell.font=WB if ci==14 else BF
        cell.fill=mf(HDRS[ci-1][1]) if ci in (14,) else mf("FEF5E7")
        cell.alignment=C if ci in (1,2,8,11,12,13,14,15,16) else W

WS={1:16,2:6,3:40,4:38,5:35,6:35,7:35,8:12,9:35,10:55,11:8,12:8,13:8,14:12,15:13,16:12}
for c,w in WS.items():
    ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes="A2"

wb.save(OUTPUT_EXCEL)
print(f"Saved: {OUTPUT_EXCEL}")
print(f"\nLevel 3 complete!")
