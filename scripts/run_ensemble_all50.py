"""
PGAI Group 5 - Ensemble on All 50 Questions
=============================================
Addresses overfitting concern raised by Prof.
- Runs ensemble config on all 50 questions
- 12 already done (loads from checkpoint)
- Runs remaining 38 fresh
- Shows aggregate delta vs Phase 2 baseline

Run from rag-skeleton directory:
  python ensemble_all50.py
"""

import re, json, requests, time, os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ── CONFIG ──
TOGETHER_API_KEY = ""
OPENAI_API_KEY   = ""
RAG_API_URL      = "http://localhost:7860"
QUESTIONS_PATH   = r"C:\Users\DELL\OneDrive - Indian Institute of Management\Desktop\Term 4\P&GAI\End Term Project\Questions_PGAI_Group_5_FINAL.xlsx"
RAG_EVAL_PATH    = "rag_phase2_evaluation_checkpoint.csv"
EXISTING_CHECKPOINT = "real_ensemble_checkpoint.csv"
INPUT_SHEET      = "All Questions"
CHECKPOINT_CSV   = "ensemble_all50_checkpoint.csv"
OUTPUT_EXCEL     = "ensemble_all50_results.xlsx"

EMBEDDING_MODEL  = "MiniLM-L6 (fast, 384d)"
ANSWER_MODEL     = "gpt-oss-20b (Together AI)"
DELAY            = 1.5

# T4 sub-questions for decomposition
T4_SUBQUESTIONS = {
    "BANKING-T4-02": [
        "What RBI rules govern LSP role in digital loan disbursement and repayment?",
        "What are the disclosure requirements for EIR calculation including upfront fees?",
        "What are the KFS delivery timelines and borrower consent rules in digital lending?"
    ],
    "BANKING-T4-04": [
        "What NBFC concentration limits apply to infrastructure project loans under RBI?",
        "What steps must an NBFC take when a loan exceeds DCCO due to court proceedings?",
        "What classification rules apply when NBFC breaches concentration norms?"
    ],
    "BANKING-T4-05": [
        "How does RBI define NPA classification for funded facilities with same borrower?",
        "When does an unpaid derivative receivable trigger NPA for other facilities?",
        "What provisioning requirements apply when NBFC classifies accounts as NPA?"
    ],
    "BANKING-T4-06": [
        "What RBI timeline applies for NBFC identification of SMA and NPA accounts?",
        "What corrective actions must NBFC take after discovering delayed NPA identification?",
        "What reporting obligations apply to NBFC under PCA framework for asset quality?"
    ],
    "BANKING-T4-09": [
        "What are the RBI rules for PCFC repayment timelines for exporters?",
        "How must a bank handle PCFC when export proceeds are delayed beyond due date?",
        "What NPA classification rules apply to overdue PCFC accounts under RBI guidelines?"
    ],
}

# ── CLIENTS ──
together_client = OpenAI(api_key=TOGETHER_API_KEY, base_url="https://api.together.xyz/v1")
nano_client     = OpenAI(api_key=OPENAI_API_KEY)
gemma_client    = OpenAI(api_key=TOGETHER_API_KEY, base_url="https://api.together.xyz/v1")

# ── ENCODING FIX ──
def fix_text(val):
    if not isinstance(val, str):
        return val
    try:
        val = val.encode('latin-1').decode('utf-8')
    except:
        pass
    val = val.replace('\u202f',' ').replace('\xa0',' ')
    val = val.replace('\xad','-').replace('\u2011','-')
    val = re.sub(r'\*\*(.*?)\*\*', r'\1', val)
    return val

# ── RETRIEVE ──
def retrieve_chunks(question, hw, top_k):
    try:
        resp = requests.post(f"{RAG_API_URL}/query", json={
            "question": question,
            "embedding_model": EMBEDDING_MODEL,
            "answer_model":    ANSWER_MODEL,
            "top_k":           top_k,
            "hybrid_weight":   hw,
        }, timeout=90)
        return resp.json().get("chunks", [])
    except Exception as e:
        print(f"      Retrieval error: {e}")
        return []

# ── DEDUPLICATE ──
def deduplicate(chunks):
    seen, unique = set(), []
    for c in chunks:
        txt = c.get("text", c.get("content",""))
        if txt and txt not in seen:
            seen.add(txt); unique.append(c)
    return unique

# ── GENERATE ──
def generate_answer(question, unique_chunks):
    if not unique_chunks:
        return ""
    context = "\n\n".join(
        f"[{i+1}] {c.get('text', c.get('content',''))}"
        for i, c in enumerate(unique_chunks)
    )
    prompt = (
        "Answer the question using only the context below. "
        "If the context does not contain the answer, say so explicitly. "
        "Cite sources inline like [1], [2].\n\n"
        f"Context:\n{context}\n\nQuestion: {question}"
    )
    try:
        resp = together_client.chat.completions.create(
            model="openai/gpt-oss-20b",
            messages=[{"role":"user","content":prompt}],
            max_completion_tokens=3000,
        )
        return fix_text(resp.choices[0].message.content or "")
    except Exception as e:
        print(f"      Generation error: {e}")
        return ""

# ── SCORE ──
RUBRIC = (
    "Rubric (max 8 pts): "
    "Factual Accuracy (0-4): 4=No errors, 3=Minor omissions, 2=Material error, 1=Hallucinated, 0=No answer. "
    "Completeness (0-2): 2=All aspects, 1=Core only, 0=Superficial. "
    "Calibration (0-2): 2=Appropriately confident, 1=Over/under confident, 0=Confident hallucination. "
    'Return JSON: {"factual_accuracy":N,"completeness":N,"confidence_calibration":N,"total":N,"reasoning":"one sentence"}'
)

PROFILES = [
    ("E1", "Strict RBI Examiner. Zero tolerance for deviation from exact RBI text. " + RUBRIC),
    ("E2", "Balanced Compliance Analyst. Fair partial credit for directionally correct answers. " + RUBRIC),
    ("E3", "Practical Compliance Officer. Would this answer lead to correct regulatory decision? " + RUBRIC),
]

SCORERS = [
    {"name":"gpt-5-nano",   "model":"gpt-5-nano",            "client":nano_client,  "reasoning":True},
    {"name":"Gemma-3N-E4B", "model":"google/gemma-3n-E4B-it","client":gemma_client,"reasoning":False},
]

def score_response(question, expected, response):
    profile_scores = {}
    for eid, system in PROFILES:
        ext = []
        for scorer in SCORERS:
            params = {
                "model":    scorer["model"],
                "messages": [
                    {"role":"system","content":system},
                    {"role":"user","content":f"QUESTION:\n{question}\n\nEXPECTED:\n{expected}\n\nRESPONSE:\n{response}\n\nScore this response."}
                ],
            }
            if scorer["reasoning"]:
                params["max_completion_tokens"] = 5000
            else:
                params["max_tokens"] = 512
                params["temperature"] = 0
            try:
                resp  = scorer["client"].chat.completions.create(**params)
                raw   = resp.choices[0].message.content or ""
                match = re.search(r'\{[^{}]+\}', raw, re.DOTALL)
                if match:
                    r     = json.loads(match.group())
                    total = (max(0,min(4,int(r.get("factual_accuracy",0)))) +
                             max(0,min(2,int(r.get("completeness",0)))) +
                             max(0,min(2,int(r.get("confidence_calibration",0)))))
                    ext.append(total)
            except Exception as e:
                print(f"         scorer error: {str(e)[:60]}")
            time.sleep(DELAY)
        if ext:
            profile_scores[eid] = round(sum(ext)/len(ext), 2)
    if profile_scores:
        return round(sum(profile_scores.values())/len(profile_scores), 2), profile_scores
    return None, {}

# ── MAIN ──
print("=" * 65)
print("  PGAI GROUP 5 - ENSEMBLE ON ALL 50 QUESTIONS")
print("=" * 65)
print(f"\n  Addressing overfitting concern: parameter validation")
print(f"  on full 50-question set (not just 12 error-recovery Qs)\n")

# Check server
print("Checking RAG server...")
try:
    requests.get(f"{RAG_API_URL}/answer-models", timeout=5)
    print("Server ready\n")
except Exception as e:
    print(f"Server not reachable: {e}"); exit(1)

# ── LOAD PHASE 2 BASELINE SCORES ──
print("Loading Phase 2 baseline scores...")
p2_df = pd.read_csv(RAG_EVAL_PATH, encoding="utf-8-sig")
p2_df.columns = [c.strip() for c in p2_df.columns]
p2_scores = dict(zip(p2_df["question_id"],
                     pd.to_numeric(p2_df["final_avg"], errors="coerce")))
print(f"Loaded {len(p2_scores)} Phase 2 scores\n")

# ── LOAD EXISTING ENSEMBLE RESULTS (12 Qs) ──
existing_results = {}
if os.path.exists(EXISTING_CHECKPOINT):
    ex = pd.read_csv(EXISTING_CHECKPOINT, encoding="utf-8-sig")
    for _, row in ex.iterrows():
        existing_results[row["question_id"]] = {
            "question_id":   row["question_id"],
            "tier":          row["tier"],
            "question_text": row.get("question_text",""),
            "expected_answer": row.get("expected_answer",""),
            "ensemble_answer": row.get("ensemble_answer",""),
            "E1":  row.get("E1",""),
            "E2":  row.get("E2",""),
            "E3":  row.get("E3",""),
            "final_score":    row.get("final_score",""),
            "baseline_score": p2_scores.get(row["question_id"],""),
            "sources":        row.get("sources",""),
        }
    print(f"Loaded {len(existing_results)} existing ensemble results (12 failing Qs)\n")

# ── LOAD ALL 50 QUESTIONS ──
print("Loading all 50 questions...")
df_all = pd.read_excel(QUESTIONS_PATH, sheet_name=INPUT_SHEET)
df_all.columns = [c.strip() for c in df_all.columns]
df_all = df_all[df_all["question_text"].notna()].reset_index(drop=True)
total  = len(df_all)
print(f"{total} questions loaded\n")

# ── LOAD CHECKPOINT ──
done_ids, all_results = set(), []
if os.path.exists(CHECKPOINT_CSV):
    ck = pd.read_csv(CHECKPOINT_CSV, encoding="utf-8-sig")
    done_ids    = set(ck["question_id"].tolist())
    all_results = ck.to_dict("records")
    print(f"Resuming checkpoint: {len(done_ids)}/{total} already done\n")
else:
    # Pre-populate with existing 12 ensemble results
    for q_id, res in existing_results.items():
        all_results.append(res)
        done_ids.add(q_id)
    print(f"Pre-loaded {len(done_ids)} existing ensemble results\n")

start_time = time.time()
remaining_qs = df_all[~df_all["question_id"].isin(done_ids)]
print(f"Remaining to run: {len(remaining_qs)} questions\n")

for i, (_, row) in enumerate(remaining_qs.iterrows()):
    q_id     = row["question_id"]
    tier     = row["tier"]
    q_text   = str(row["question_text"]).strip()
    expected = str(row.get("expected_answer","")).strip()
    baseline = p2_scores.get(q_id,"")
    is_t4    = "T4" in q_id

    print(f"\n[{i+1}/{len(remaining_qs)}] {q_id} | Tier {tier} | Baseline: {baseline}")

    # ── Retrieve with 3 retrievers ──
    all_chunks = []
    for hw, k, label in [(0.0,6,"R1"),(0.5,4,"R2"),(1.0,6,"R3")]:
        chunks = retrieve_chunks(q_text, hw, k)
        print(f"   {label} (hw={hw},K={k}): {len(chunks)} chunks")
        all_chunks.extend(chunks)
        time.sleep(0.5)

    # ── Sub-Q retrieval for T4 ──
    if is_t4 and q_id in T4_SUBQUESTIONS:
        for j, sq in enumerate(T4_SUBQUESTIONS[q_id]):
            chunks = retrieve_chunks(sq, 0.0, 6)
            print(f"   Sub-Q{j+1}: {len(chunks)} chunks")
            all_chunks.extend(chunks)
            time.sleep(0.5)

    # ── Deduplicate ──
    unique  = deduplicate(all_chunks)
    sources = "; ".join(set(
        c.get("filename","?").split("\\")[-1].split("/")[-1]
        for c in unique[:8]
    ))
    print(f"   Unique chunks: {len(unique)}")

    # ── Generate ──
    answer = generate_answer(q_text, unique)
    if not answer:
        print(f"   Generation failed — skipping")
        continue
    print(f"   Answer: {answer[:80]}...")

    # ── Score ──
    print(f"   Scoring...")
    final_score, ps = score_response(q_text, expected, answer)
    improvement = round(final_score - float(baseline), 2) if final_score and baseline else ""

    print(f"   Score: {final_score}/8 | Baseline: {baseline} | Delta: {improvement}")

    all_results.append({
        "question_id":     q_id,
        "tier":            tier,
        "question_text":   q_text,
        "expected_answer": expected,
        "ensemble_answer": answer,
        "E1":              ps.get("E1",""),
        "E2":              ps.get("E2",""),
        "E3":              ps.get("E3",""),
        "final_score":     final_score or "",
        "baseline_score":  baseline,
        "improvement":     improvement,
        "sources":         sources,
    })

    pd.DataFrame(all_results).to_csv(CHECKPOINT_CSV, index=False, encoding="utf-8-sig")
    elapsed   = time.time() - start_time
    remaining = (elapsed/(i+1))*(len(remaining_qs)-i-1) if i+1 else 0
    print(f"   Saved | {elapsed/60:.1f}m | ~{remaining/60:.1f}m remaining")

results_df = pd.DataFrame(all_results)
results_df["final_score"]    = pd.to_numeric(results_df["final_score"],   errors="coerce")
results_df["baseline_score"] = pd.to_numeric(results_df["baseline_score"],errors="coerce")
results_df["improvement"]    = pd.to_numeric(results_df["improvement"],   errors="coerce")

# ── SUMMARY ──
print(f"\n{'='*65}")
print("  ENSEMBLE ALL 50 - COMPLETE RESULTS")
print(f"{'='*65}")

overall_base = results_df["baseline_score"].mean()
overall_ens  = results_df["final_score"].mean()
improved     = (results_df["improvement"] > 0).sum()
declined     = (results_df["improvement"] < 0).sum()
same         = (results_df["improvement"] == 0).sum()

print(f"\n  Phase 2 Baseline (all 50) : {overall_base:.2f}/8")
print(f"  Ensemble (all 50)         : {overall_ens:.2f}/8")
print(f"  Overall Delta             : {overall_ens-overall_base:+.2f}")
print(f"  Improved                  : {improved}/50")
print(f"  Declined                  : {declined}/50")
print(f"  No change                 : {same}/50")

print(f"\n  By Tier:")
for tier in sorted(results_df["tier"].dropna().unique()):
    t  = results_df[results_df["tier"]==tier]
    b  = t["baseline_score"].mean()
    e  = t["final_score"].mean()
    im = (t["improvement"] > 0).sum()
    dc = (t["improvement"] < 0).sum()
    print(f"    T{int(tier)}: Baseline={b:.2f} | Ensemble={e:.2f} | Delta={e-b:+.2f} | Improved={im}/10 | Declined={dc}/10")

# ── Recovery vs Non-recovery split ──
FAILING_IDS = ["BANKING-T2-03","BANKING-T2-10","BANKING-T3-04",
               "BANKING-T4-02","BANKING-T4-04","BANKING-T4-05",
               "BANKING-T4-06","BANKING-T4-09","BANKING-T5-01",
               "BANKING-T5-02","BANKING-T5-06","BANKING-T5-10"]

error_set  = results_df[results_df["question_id"].isin(FAILING_IDS)]
holdout    = results_df[~results_df["question_id"].isin(FAILING_IDS)]

print(f"\n  Error Recovery Set (12 Qs):")
print(f"    Baseline: {error_set['baseline_score'].mean():.2f}/8")
print(f"    Ensemble: {error_set['final_score'].mean():.2f}/8")
print(f"    Delta   : {error_set['final_score'].mean()-error_set['baseline_score'].mean():+.2f}")
print(f"    Improved: {(error_set['improvement']>0).sum()}/12 | Declined: {(error_set['improvement']<0).sum()}/12")

print(f"\n  Natural Hold-out (38 Qs, scored 6+):")
print(f"    Baseline: {holdout['baseline_score'].mean():.2f}/8")
print(f"    Ensemble: {holdout['final_score'].mean():.2f}/8")
print(f"    Delta   : {holdout['final_score'].mean()-holdout['baseline_score'].mean():+.2f}")
print(f"    Improved: {(holdout['improvement']>0).sum()}/38 | Declined: {(holdout['improvement']<0).sum()}/38")

# ── WRITE EXCEL ──
print(f"\nWriting: {OUTPUT_EXCEL}")
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "All 50 Ensemble Results"

HF  = Font(name="Arial",bold=True,color="FFFFFF",size=10)
BF  = Font(name="Arial",size=10)
WBF = Font(name="Arial",bold=True,color="FFFFFF",size=11)
C   = Alignment(horizontal="center",vertical="center",wrap_text=True)
W   = Alignment(wrap_text=True,vertical="top")
TIER_CLR = {"1":"EBF5FB","2":"E9F7EF","3":"FEF9E7","4":"FEF5E7","5":"FDEDEC"}
def mf(h): return PatternFill("solid",fgColor=h)

HEADERS = [
    ("question_id","1F4E79"),("tier","1F4E79"),
    ("question_text","1F4E79"),("expected_answer","1F4E79"),
    ("ensemble_answer","1A5276"),("sources","1A5276"),
    ("E1","117A65"),("E2","117A65"),("E3","117A65"),
    ("Ensemble Score","117A65"),("Phase2 Baseline","1F4E79"),
    ("Delta","1F4E79"),("Error Recovery Set","1F4E79"),
]

for c,(h,col) in enumerate(HEADERS,start=1):
    cell=ws.cell(row=1,column=c,value=h)
    cell.font=HF; cell.fill=mf(col); cell.alignment=C
ws.row_dimensions[1].height=35

for ri,row in results_df.iterrows():
    er    = ri+2
    t_str = str(row.get("tier","")).strip()
    t_fill= mf(TIER_CLR.get(t_str,"FFFFFF"))
    imp   = row.get("improvement","")
    is_er = row.get("question_id","") in FAILING_IDS

    vals=[
        row.get("question_id",""), row.get("tier",""),
        row.get("question_text",""), row.get("expected_answer",""),
        row.get("ensemble_answer",""), row.get("sources",""),
        row.get("E1",""), row.get("E2",""), row.get("E3",""),
        row.get("final_score",""), row.get("baseline_score",""),
        imp, "Yes" if is_er else "No",
    ]
    for ci,val in enumerate(vals,start=1):
        val  = "" if pd.isna(val) else val
        cell = ws.cell(row=er,column=ci,value=val)
        if ci==10:
            cell.font=WBF; cell.fill=mf("117A65")
        elif ci==12:
            try:
                v = float(val)
                cell.fill = mf("D5F5E3") if v>0 else mf("FADBD8") if v<0 else t_fill
                cell.font = Font(name="Arial",bold=True,size=10)
            except:
                cell.font=BF; cell.fill=t_fill
        else:
            cell.font=BF; cell.fill=t_fill
        cell.alignment = C if ci not in (3,4,5,6) else W

WIDTHS={1:18,2:6,3:40,4:35,5:50,6:35,7:8,8:8,9:8,10:12,11:14,12:10,13:15}
for c,w in WIDTHS.items():
    ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes="A2"

# ── Summary tab ──
ws2 = wb.create_sheet("Generalisation Summary")
sum_rows = [
    ["","Phase 2 Baseline","Ensemble","Delta","Improved","Declined"],
    ["Error Recovery Set (12 Qs)",
     round(error_set["baseline_score"].mean(),2),
     round(error_set["final_score"].mean(),2),
     round(error_set["final_score"].mean()-error_set["baseline_score"].mean(),2),
     f"{(error_set['improvement']>0).sum()}/12",
     f"{(error_set['improvement']<0).sum()}/12"],
    ["Natural Hold-out (38 Qs)",
     round(holdout["baseline_score"].mean(),2),
     round(holdout["final_score"].mean(),2),
     round(holdout["final_score"].mean()-holdout["baseline_score"].mean(),2),
     f"{(holdout['improvement']>0).sum()}/38",
     f"{(holdout['improvement']<0).sum()}/38"],
    ["ALL 50 QUESTIONS",
     round(overall_base,2),
     round(overall_ens,2),
     round(overall_ens-overall_base,2),
     f"{improved}/50",
     f"{declined}/50"],
]

for ri,row in enumerate(sum_rows,start=1):
    for ci,val in enumerate(row,start=1):
        cell=ws2.cell(row=ri,column=ci,value=val)
        if ri==1:
            cell.font=HF; cell.fill=mf("1F4E79")
        elif ri==4:
            cell.font=Font(name="Arial",bold=True,size=10)
            cell.fill=mf("D5E8F3")
        else:
            cell.font=BF
            cell.fill=mf("E9F7EF") if ri==2 else mf("EBF5FB")
        cell.alignment=C
for c,w in {1:28,2:16,3:12,4:12,5:14,6:14}.items():
    ws2.column_dimensions[get_column_letter(c)].width=w

wb.save(OUTPUT_EXCEL)
print(f"Saved: {OUTPUT_EXCEL}")
print(f"\nGeneralisation Summary tab shows error recovery vs hold-out split.")
print(f"Use this to address Prof's overfitting concern.")
