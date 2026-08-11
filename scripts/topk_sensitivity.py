"""
PGAI Group 5 — Top-K Sensitivity Analysis (Level 2)
=====================================================
Tests Top-K = 2, 4, 6, 8 on 12 failing questions
with current chunk size = 800 (no re-ingestion needed)

Run from rag-skeleton directory:
  python topk_sensitivity.py
"""

import re
import json
import requests
import time
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ============================================================
# CONFIGURATION
# ============================================================
TOGETHER_API_KEY = ""
OPENAI_API_KEY   = ""
RAG_API_URL      = "http://localhost:7860"
QUESTIONS_PATH   = r"C:\Users\DELL\OneDrive - Indian Institute of Management\Desktop\Term 4\P&GAI\End Term Project\Questions_PGAI_Group_5_FINAL.xlsx"
INPUT_SHEET      = "All Questions"
OUTPUT_EXCEL     = "topk_sensitivity_results.xlsx"
CHECKPOINT_CSV   = "topk_sensitivity_checkpoint.csv"

EMBEDDING_MODEL  = "MiniLM-L6 (fast, 384d)"
ANSWER_MODEL     = "gpt-oss-20b (Together AI)"
HYBRID_WEIGHT    = 0.5   # Keep constant — testing Top-K only
DELAY            = 1.5

# 12 failing question IDs
FAILING_IDS = [
    "BANKING-T2-03", "BANKING-T2-10",
    "BANKING-T3-04",
    "BANKING-T4-02", "BANKING-T4-04", "BANKING-T4-05",
    "BANKING-T4-06", "BANKING-T4-09",
    "BANKING-T5-01", "BANKING-T5-02", "BANKING-T5-06", "BANKING-T5-10",
]

# Top-K values to test
TOPK_CONFIGS = [
    {"top_k": 2,  "label": "Top-K=2 (narrow)"},
    {"top_k": 4,  "label": "Top-K=4 (baseline)"},
    {"top_k": 6,  "label": "Top-K=6 (wider)"},
    {"top_k": 8,  "label": "Top-K=8 (broadest)"},
]

# Phase 2 baseline scores
BASELINE_SCORES = {
    "BANKING-T2-03": 5.17, "BANKING-T2-10": 4.83,
    "BANKING-T3-04": 5.50,
    "BANKING-T4-02": 1.83, "BANKING-T4-04": 5.50,
    "BANKING-T4-05": 5.33, "BANKING-T4-06": 5.67,
    "BANKING-T4-09": 5.50,
    "BANKING-T5-01": 5.17, "BANKING-T5-02": 5.67,
    "BANKING-T5-06": 5.50, "BANKING-T5-10": 5.17,
}

# ============================================================
# ENCODING FIX
# ============================================================
def fix_response(val):
    if not isinstance(val, str):
        return val
    try:
        val = val.encode('latin-1').decode('utf-8')
    except:
        pass
    val = val.replace('\u202f', ' ')
    val = val.replace('\xa0', ' ')
    val = val.replace('\xad', '-')
    val = val.replace('\u2011', '-')
    val = val.replace('【', '[').replace('】', ']')
    val = re.sub(r'\*\*(.*?)\*\*', r'\1', val)
    val = re.sub(r'\*(.*?)\*', r'\1', val)
    return val

# ============================================================
# EVALUATOR SETUP — same as Phase 2
# ============================================================
EVALUATORS_CFG = [
    {
        "name": "gpt-5-nano",
        "model": "gpt-5-nano",
        "base_url": None,
        "api_key": OPENAI_API_KEY,
        "reasoning": True,
    },
    {
        "name": "Gemma-3N-E4B",
        "model": "google/gemma-3n-E4B-it",
        "base_url": "https://api.together.xyz/v1",
        "api_key": TOGETHER_API_KEY,
        "reasoning": False,
    },
]

EVAL_CLIENTS = {}
for cfg in EVALUATORS_CFG:
    if cfg["base_url"]:
        EVAL_CLIENTS[cfg["name"]] = OpenAI(
            api_key=cfg["api_key"], base_url=cfg["base_url"])
    else:
        EVAL_CLIENTS[cfg["name"]] = OpenAI(api_key=cfg["api_key"])

EVAL_PROFILES = [
    {
        "id": "E1", "label": "Strict RBI Examiner",
        "system": """You are a strict RBI regulatory examiner scoring AI responses.
Rubric (max 8 pts):
Factual Accuracy (0-4): 4=No errors|3=Minor omissions|2=Material error|1=Hallucinated|0=No answer
Completeness (0-2): 2=All aspects|1=Core only|0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident|1=Over/under|0=Confident hallucination
Respond ONLY in JSON: {"factual_accuracy":<0-4>,"completeness":<0-2>,"confidence_calibration":<0-2>,"total":<sum>,"reasoning":"<one sentence>"}"""
    },
    {
        "id": "E2", "label": "Balanced Compliance Analyst",
        "system": """You are a balanced compliance analyst scoring AI responses.
Rubric (max 8 pts):
Factual Accuracy (0-4): 4=No errors|3=Minor omissions|2=Material error|1=Hallucinated|0=No answer
Completeness (0-2): 2=All aspects|1=Core only|0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident|1=Over/under|0=Confident hallucination
Respond ONLY in JSON: {"factual_accuracy":<0-4>,"completeness":<0-2>,"confidence_calibration":<0-2>,"total":<sum>,"reasoning":"<one sentence>"}"""
    },
    {
        "id": "E3", "label": "Practical Compliance Officer",
        "system": """You are a senior compliance officer scoring AI responses for real-world usefulness.
Rubric (max 8 pts):
Factual Accuracy (0-4): 4=No errors|3=Minor omissions|2=Material error|1=Hallucinated|0=No answer
Completeness (0-2): 2=All aspects|1=Core only|0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident|1=Over/under|0=Confident hallucination
Respond ONLY in JSON: {"factual_accuracy":<0-4>,"completeness":<0-2>,"confidence_calibration":<0-2>,"total":<sum>,"reasoning":"<one sentence>"}"""
    },
]

def score_response(question, expected, response):
    profile_scores = {}
    for ev in EVAL_PROFILES:
        ext_scores = []
        for cfg in EVALUATORS_CFG:
            client = EVAL_CLIENTS[cfg["name"]]
            user_msg = f"QUESTION:\n{question}\n\nEXPECTED ANSWER:\n{expected}\n\nRAG RESPONSE:\n{response}\n\nScore using the rubric."
            params = {
                "model": cfg["model"],
                "messages": [
                    {"role": "system", "content": ev["system"]},
                    {"role": "user",   "content": user_msg}
                ],
            }
            if cfg["reasoning"]:
                params["max_completion_tokens"] = 5000
            else:
                params["max_tokens"]  = 512
                params["temperature"] = 0
            try:
                resp  = client.chat.completions.create(**params)
                raw   = resp.choices[0].message.content or ""
                match = re.search(r'\{.*?\}', raw, re.DOTALL)
                if match:
                    result = json.loads(match.group())
                    total  = (max(0,min(4,int(result.get("factual_accuracy",0)))) +
                              max(0,min(2,int(result.get("completeness",0)))) +
                              max(0,min(2,int(result.get("confidence_calibration",0)))))
                    ext_scores.append(total)
            except Exception as e:
                print(f"         scorer error: {str(e)[:60]}")
            time.sleep(DELAY)
        if ext_scores:
            profile_scores[ev["id"]] = round(sum(ext_scores)/len(ext_scores), 2)
    if profile_scores:
        return round(sum(profile_scores.values())/len(profile_scores), 2), profile_scores
    return None, {}

# ============================================================
# STEP 1: Verify server
# ============================================================
print("=" * 65)
print("  PGAI GROUP 5 — TOP-K SENSITIVITY ANALYSIS (Level 2)")
print("=" * 65)
print(f"\n  12 questions × 4 Top-K configs = 48 RAG calls")
print(f"  Scoring calls: 48 × 6 = 288")
print(f"  Hybrid weight fixed at: {HYBRID_WEIGHT}")
print(f"  Chunk size fixed at   : 800 (no re-ingestion needed)\n")

print("🔍 Checking RAG server...")
try:
    r = requests.get(f"{RAG_API_URL}/answer-models", timeout=5)
    print(f"✅ Server ready\n")
except Exception as e:
    print(f"❌ Server not reachable: {e}")
    exit(1)

# ============================================================
# STEP 2: Load questions
# ============================================================
print("📂 Loading questions...")
df_all  = pd.read_excel(QUESTIONS_PATH, sheet_name=INPUT_SHEET)
df_all.columns = [c.strip() for c in df_all.columns]
df_fail = df_all[df_all["question_id"].isin(FAILING_IDS)].reset_index(drop=True)
print(f"✅ {len(df_fail)} failing questions loaded\n")

# ============================================================
# STEP 3: Resume from checkpoint
# ============================================================
done_keys   = set()
all_results = []
if os.path.exists(CHECKPOINT_CSV):
    ck = pd.read_csv(CHECKPOINT_CSV, encoding="utf-8-sig")
    done_keys   = set(zip(ck["question_id"], ck["top_k"].astype(str)))
    all_results = ck.to_dict("records")
    print(f"♻️  Resuming — {len(done_keys)} already done\n")
else:
    print("🚀 Fresh start\n")

# ============================================================
# STEP 4: Run all Top-K combinations
# ============================================================
start_time = time.time()
total_runs = len(df_fail) * len(TOPK_CONFIGS)
run_num    = 0

for _, row in df_fail.iterrows():
    q_id     = row["question_id"]
    tier     = row["tier"]
    q_text   = str(row["question_text"]).strip()
    expected = str(row.get("expected_answer","")).strip()
    baseline = BASELINE_SCORES.get(q_id,"")

    for cfg in TOPK_CONFIGS:
        top_k = cfg["top_k"]
        label = cfg["label"]
        key   = (q_id, str(top_k))
        run_num += 1

        if key in done_keys:
            print(f"[{run_num}/{total_runs}] {q_id} | top_k={top_k} ⏭ skipped")
            continue

        print(f"\n[{run_num}/{total_runs}] {q_id} | Tier {tier} | {label}")

        # RAG call
        payload = {
            "question":        q_text,
            "embedding_model": EMBEDDING_MODEL,
            "answer_model":    ANSWER_MODEL,
            "top_k":           top_k,
            "hybrid_weight":   HYBRID_WEIGHT,
        }
        try:
            resp    = requests.post(f"{RAG_API_URL}/query", json=payload, timeout=60)
            data    = resp.json()
            rag_ans = fix_response(data.get("answer",""))
            chunks  = data.get("chunks",[])
            top_src = "; ".join([f"{c.get('filename','?')} ({c.get('score',0):.3f})"
                                  for c in chunks[:3]])
            top_score = chunks[0].get("score",0) if chunks else 0
            rag_status = "OK" if rag_ans else "EMPTY"
        except Exception as e:
            rag_ans    = ""
            top_src    = ""
            top_score  = 0
            rag_status = f"ERROR: {str(e)[:60]}"

        print(f"   RAG: {rag_status} | top_score={top_score:.3f}")

        # Score
        if rag_ans:
            final_score, profile_scores = score_response(q_text, expected, rag_ans)
        else:
            final_score    = None
            profile_scores = {}

        improvement = round(final_score - baseline, 2) if final_score and baseline else ""

        all_results.append({
            "question_id":    q_id,
            "tier":           tier,
            "question_text":  q_text,
            "top_k":          top_k,
            "config_label":   label,
            "hybrid_weight":  HYBRID_WEIGHT,
            "rag_response":   rag_ans,
            "top_sources":    top_src,
            "top_similarity": round(top_score,3),
            "E1":             profile_scores.get("E1",""),
            "E2":             profile_scores.get("E2",""),
            "E3":             profile_scores.get("E3",""),
            "final_score":    final_score if final_score else "",
            "baseline_score": baseline,
            "improvement":    improvement,
            "rag_status":     rag_status,
        })

        pd.DataFrame(all_results).to_csv(
            CHECKPOINT_CSV, index=False, encoding="utf-8-sig")

        elapsed   = time.time() - start_time
        remaining = (elapsed/run_num)*(total_runs-run_num) if run_num else 0
        print(f"   Score: {final_score}/8 (baseline: {baseline}, Δ {improvement})")
        print(f"   💾 | ⏱ {elapsed/60:.1f}m | ~{remaining/60:.1f}m remaining")

results_df = pd.DataFrame(all_results)

# ============================================================
# SUMMARY
# ============================================================
print(f"\n{'='*65}")
print("  TOP-K SENSITIVITY — SUMMARY")
print(f"{'='*65}")
for cfg in TOPK_CONFIGS:
    tk  = cfg["top_k"]
    sub = results_df[results_df["top_k"]==tk]
    avg = pd.to_numeric(sub["final_score"],errors="coerce").mean()
    base= pd.to_numeric(sub["baseline_score"],errors="coerce").mean()
    print(f"\n  {cfg['label']}")
    print(f"    Avg score    : {avg:.2f}/8")
    print(f"    Avg Δ        : {avg-base:+.2f}")
print(f"\n{'='*65}\n")

# ============================================================
# WRITE EXCEL WITH HEATMAP
# ============================================================
print(f"💾 Writing: {OUTPUT_EXCEL}")
wb  = openpyxl.Workbook()

# ── Tab 1: Full Results ──
ws1 = wb.active
ws1.title = "Top-K Results"

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
WHITE_BOLD= Font(name="Arial", bold=True, color="FFFFFF", size=11)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP      = Alignment(wrap_text=True, vertical="top")
TIER_CLR  = {"1":"EBF5FB","2":"E9F7EF","3":"FEF9E7","4":"FEF5E7","5":"FDEDEC"}
TOPK_CLR  = {"2":"F9EBEA","4":"D6EAF8","6":"D5F5E3","8":"FEF9E7"}
TOPK_HDR  = {"2":"922B21","4":"1A5276","6":"1E8449","8":"7D6608"}
def mf(h): return PatternFill("solid", fgColor=h)

HEADERS = [
    ("question_id","1F4E79"),("tier","1F4E79"),
    ("question_text","1F4E79"),("config_label","1F4E79"),
    ("top_k","1F4E79"),("rag_response","1F4E79"),
    ("top_sources","1F4E79"),("top_similarity","1F4E79"),
    ("E1","1F4E79"),("E2","1F4E79"),("E3","1F4E79"),
    ("final_score","1F4E79"),("baseline_score","1F4E79"),
    ("improvement","1F4E79"),
]
for c,(label,color) in enumerate(HEADERS,start=1):
    cell=ws1.cell(row=1,column=c,value=label)
    cell.font=HDR_FONT; cell.fill=mf(color); cell.alignment=CENTER
ws1.row_dimensions[1].height=30

for ri,row in results_df.iterrows():
    er   = ri+2
    tk   = str(int(row.get("top_k",4)))
    fill = mf(TOPK_CLR.get(tk,"FFFFFF"))
    data_row = [
        row.get("question_id",""), row.get("tier",""),
        row.get("question_text",""), row.get("config_label",""),
        row.get("top_k",""), row.get("rag_response",""),
        row.get("top_sources",""), row.get("top_similarity",""),
        row.get("E1",""), row.get("E2",""), row.get("E3",""),
        row.get("final_score",""), row.get("baseline_score",""),
        row.get("improvement",""),
    ]
    for ci,val in enumerate(data_row,start=1):
        val  = "" if pd.isna(val) else val
        cell = ws1.cell(row=er,column=ci,value=val)
        is_score = ci in (9,10,11,12,13,14)
        cell.font = WHITE_BOLD if ci==12 else BOLD_FONT if is_score else BODY_FONT
        cell.fill = mf(TOPK_HDR.get(tk,"1F4E79")) if ci==12 else fill
        cell.alignment = CENTER if ci not in (3,6,7) else WRAP

WIDTHS={1:16,2:6,3:42,4:18,5:8,6:55,7:40,8:13,
        9:8,10:8,11:8,12:12,13:13,14:12}
for c,w in WIDTHS.items():
    ws1.column_dimensions[get_column_letter(c)].width=w
ws1.freeze_panes="A2"

# ── Tab 2: Heatmap ──
ws2 = wb.create_sheet("Heatmap")
ws2.cell(row=1,column=1,value="Question ID").font = HDR_FONT
ws2.cell(row=1,column=1).fill = mf("1F4E79")
ws2.cell(row=1,column=2,value="Tier").font = HDR_FONT
ws2.cell(row=1,column=2).fill = mf("1F4E79")
ws2.cell(row=1,column=3,value="Baseline").font = HDR_FONT
ws2.cell(row=1,column=3).fill = mf("1F4E79")

for ci,cfg in enumerate(TOPK_CONFIGS,start=4):
    tk = cfg["top_k"]
    cell=ws2.cell(row=1,column=ci,value=f"Top-K={tk}")
    cell.font=HDR_FONT; cell.fill=mf(TOPK_HDR.get(str(tk),"1F4E79"))
    cell.alignment=CENTER

ws2.cell(row=1,column=len(TOPK_CONFIGS)+4,value="Best Top-K").font=HDR_FONT
ws2.cell(row=1,column=len(TOPK_CONFIGS)+4).fill=mf("1F4E79")
ws2.row_dimensions[1].height=28

def score_to_color(score, baseline):
    """Green if improved, red if declined, yellow if same."""
    try:
        s = float(score); b = float(baseline)
        if s >= b + 0.5: return "D5F5E3"   # green — improved
        elif s <= b - 0.5: return "FADBD8"  # red — declined
        else: return "FEF9E7"               # yellow — similar
    except:
        return "FFFFFF"

for ri,q_id in enumerate(FAILING_IDS, start=2):
    q_data = results_df[results_df["question_id"]==q_id]
    if q_data.empty:
        continue
    tier     = q_data.iloc[0]["tier"]
    baseline = BASELINE_SCORES.get(q_id,"")
    t_fill   = mf(TIER_CLR.get(str(int(tier)),"FFFFFF"))

    ws2.cell(row=ri,column=1,value=q_id).fill=t_fill
    ws2.cell(row=ri,column=2,value=f"T{int(tier)}").fill=t_fill
    ws2.cell(row=ri,column=3,value=baseline).fill=t_fill

    best_score = -1
    best_tk    = ""
    for ci,cfg in enumerate(TOPK_CONFIGS, start=4):
        tk  = cfg["top_k"]
        row = q_data[q_data["top_k"]==tk]
        score = row["final_score"].values[0] if not row.empty else ""
        color = score_to_color(score, baseline)
        cell  = ws2.cell(row=ri, column=ci, value=score)
        cell.fill      = mf(color)
        cell.alignment = CENTER
        cell.font      = BOLD_FONT
        try:
            if float(score) > best_score:
                best_score = float(score)
                best_tk    = f"Top-K={tk} ({score})"
        except:
            pass

    ws2.cell(row=ri,column=len(TOPK_CONFIGS)+4,value=best_tk).font=BOLD_FONT

for c,w in {1:16,2:6,3:10,4:12,5:12,6:12,7:12,8:20}.items():
    ws2.column_dimensions[get_column_letter(c)].width=w
ws2.freeze_panes="A2"

# ── Tab 3: Summary ──
ws3 = wb.create_sheet("Summary")
sum_hdrs=["Top-K Config","Avg Score /8","Avg Baseline","Avg Δ","Questions Improved","Questions Declined"]
for c,h in enumerate(sum_hdrs,start=1):
    cell=ws3.cell(row=1,column=c,value=h)
    cell.font=HDR_FONT; cell.fill=mf("1F4E79"); cell.alignment=CENTER
ws3.row_dimensions[1].height=28

for ri,cfg in enumerate(TOPK_CONFIGS,start=2):
    tk   = cfg["top_k"]
    sub  = results_df[results_df["top_k"]==tk]
    avg  = pd.to_numeric(sub["final_score"],errors="coerce").mean()
    base = pd.to_numeric(sub["baseline_score"],errors="coerce").mean()
    imp  = pd.to_numeric(sub["improvement"],errors="coerce")
    improved = (imp > 0).sum()
    declined = (imp < 0).sum()
    fill = mf(TOPK_CLR.get(str(tk),"FFFFFF"))
    vals=[cfg["label"],round(avg,2),round(base,2),round(avg-base,2),improved,declined]
    for c,v in enumerate(vals,start=1):
        cell=ws3.cell(row=ri,column=c,value=v)
        cell.font=BOLD_FONT if c in (2,4) else BODY_FONT
        cell.fill=fill; cell.alignment=CENTER
for c in range(1,len(sum_hdrs)+1):
    ws3.column_dimensions[get_column_letter(c)].width=20
ws3.freeze_panes="A2"

wb.save(OUTPUT_EXCEL)
print(f"✅ Saved: {OUTPUT_EXCEL}")
print(f"✅ Checkpoint: {CHECKPOINT_CSV}")
print(f"\n🎉 Top-K sensitivity analysis complete!")
print(f"   Open Heatmap tab — green = improved, red = declined vs baseline")
