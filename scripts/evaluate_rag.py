# ============================================================
# PGAI Group 5 — Phase 2 RAG Evaluation Script
# ============================================================
# RAG Model  : gpt-oss-20b (the model being evaluated)
# Evaluators : gpt-5-nano + Gemma (gpt-oss-20b excluded — no self-eval)
# Profiles   : E1 Strict | E2 Balanced | E3 Practical (same as Phase 1)
# Final Score: avg(E1 + E2 + E3), each = avg(nano + Gemma)
# Total calls: 50 × 3 profiles × 2 scorers = 300
# ============================================================

import json, re, time, os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openai import OpenAI

# ============================================================
# CONFIGURATION
# ============================================================
OPENAI_API_KEY   = ""
TOGETHER_API_KEY = ""

INPUT_CSV      = "rag_phase2_responses.csv"
CHECKPOINT_CSV = "rag_phase2_evaluation_checkpoint.csv"
OUTPUT_EXCEL   = "PGAI_Group5_RAG_Evaluation.xlsx"

DELAY = 1.0

# gpt-oss-20b is the RAG model — excluded from evaluation (no self-eval)
EVALUATOR_MODELS = [
    {
        "name":     "gpt-5-nano",
        "model":    "gpt-5-nano",
        "base_url": None,
        "api_key":  OPENAI_API_KEY,
        "reasoning": True,    # no temperature, use max_completion_tokens
    },
    {
        "name":     "Gemma-3N-E4B",
        "model":    "google/gemma-3n-E4B-it",
        "base_url": "https://api.together.xyz/v1",
        "api_key":  TOGETHER_API_KEY,
        "reasoning": False,   # standard model
    },
]

# ============================================================
# SAME 3 EVALUATOR PROFILES AS PHASE 1 — DO NOT CHANGE
# ============================================================
EVALUATORS = [
    {
        "id":    "E1",
        "label": "Strict RBI Examiner",
        "system": """You are a strict RBI regulatory examiner scoring AI responses to Indian banking compliance questions.

Scoring philosophy:
- Any deviation from exact RBI text, threshold, date = material error
- Global norms (FATF, Basel) cited instead of RBI-specific rules = factually wrong
- Wrong thresholds (%, amounts, timelines) = score 2 at most on factual accuracy
- Confident wrong answer = calibration score 0
- Zero tolerance for hallucination, reward precision above all

RUBRIC (max 8 pts):
Factual Accuracy (0-4): 4=No errors | 3=Minor omissions | 2=Material error | 1=Hallucinated | 0=No answer
Completeness (0-2): 2=All aspects | 1=Core only | 0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident | 1=Over/under confident | 0=Confident hallucination

Respond ONLY in this JSON format, nothing else:
{"factual_accuracy": <0-4>, "completeness": <0-2>, "confidence_calibration": <0-2>, "total": <sum>, "reasoning": "<one sentence>"}"""
    },
    {
        "id":    "E2",
        "label": "Balanced Compliance Analyst",
        "system": """You are a balanced compliance analyst scoring AI responses to Indian banking regulatory questions.

Scoring philosophy:
- Give fair partial credit for directionally correct answers
- Carefully distinguish minor omissions from material errors
- Reward correct regulatory framework identification even if threshold slightly off
- Overconfidence on wrong facts = calibration 0, no leniency
- Balance strictness with fairness

RUBRIC (max 8 pts):
Factual Accuracy (0-4): 4=No errors | 3=Minor omissions | 2=Material error | 1=Hallucinated | 0=No answer
Completeness (0-2): 2=All aspects | 1=Core only | 0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident | 1=Over/under confident | 0=Confident hallucination

Respond ONLY in this JSON format, nothing else:
{"factual_accuracy": <0-4>, "completeness": <0-2>, "confidence_calibration": <0-2>, "total": <sum>, "reasoning": "<one sentence>"}"""
    },
    {
        "id":    "E3",
        "label": "Practical Compliance Officer",
        "system": """You are a senior compliance officer at an Indian bank scoring AI responses for real-world regulatory usefulness.

Scoring philosophy:
- Ask: would a compliance officer make the correct decision using this answer?
- Right outcome through correct reasoning = high score even with minor gaps
- Wrong threshold/regulation leading to wrong decision = low score
- Appropriate hedging on uncertain points = good calibration
- Confident hallucination misleading a compliance team = heavily penalised

RUBRIC (max 8 pts):
Factual Accuracy (0-4): 4=No errors | 3=Minor omissions | 2=Material error | 1=Hallucinated | 0=No answer
Completeness (0-2): 2=All aspects | 1=Core only | 0=Superficial
Confidence Calibration (0-2): 2=Appropriately confident | 1=Over/under confident | 0=Confident hallucination

Respond ONLY in this JSON format, nothing else:
{"factual_accuracy": <0-4>, "completeness": <0-2>, "confidence_calibration": <0-2>, "total": <sum>, "reasoning": "<one sentence>"}"""
    },
]

# ============================================================
# INITIALISE CLIENTS
# ============================================================
CLIENTS = {}
for cfg in EVALUATOR_MODELS:
    if cfg["base_url"]:
        CLIENTS[cfg["name"]] = OpenAI(api_key=cfg["api_key"], base_url=cfg["base_url"])
    else:
        CLIENTS[cfg["name"]] = OpenAI(api_key=cfg["api_key"])

# ============================================================
# SCORING FUNCTION
# ============================================================
def score_one(scorer_cfg, evaluator, question, expected, response):
    client = CLIENTS[scorer_cfg["name"]]
    user_msg = f"""QUESTION:
{question}

EXPECTED ANSWER (ground truth):
{expected}

RAG SYSTEM RESPONSE TO EVALUATE:
{response}

Score this response using the rubric."""

    params = {
        "model":    scorer_cfg["model"],
        "messages": [
            {"role": "system", "content": evaluator["system"]},
            {"role": "user",   "content": user_msg}
        ],
    }
    # reasoning models: max_completion_tokens, no temperature
    if scorer_cfg["reasoning"]:
        params["max_completion_tokens"] = 5000
    else:
        params["max_tokens"]  = 512
        params["temperature"] = 0

    try:
        resp  = client.chat.completions.create(**params)
        raw   = resp.choices[0].message.content
        if not raw or not raw.strip():
            return None, "EMPTY"
        match = re.search(r'\{.*?\}', raw, re.DOTALL)
        if not match:
            return None, "PARSE_ERROR"
        result = json.loads(match.group())
        fa  = max(0, min(4, int(result.get("factual_accuracy", 0))))
        co  = max(0, min(2, int(result.get("completeness", 0))))
        cc  = max(0, min(2, int(result.get("confidence_calibration", 0))))
        return {
            "factual": fa, "completeness": co,
            "calibration": cc, "total": fa+co+cc,
            "reasoning": result.get("reasoning","")
        }, "OK"
    except Exception as e:
        return None, f"ERROR: {str(e)[:80]}"

# ============================================================
# STEP 1: Load responses
# ============================================================
print("=" * 60)
print("  PGAI GROUP 5 — RAG EVALUATION (Phase 2)")
print("=" * 60)
print(f"\nRAG Model     : gpt-oss-20b (being evaluated)")
print(f"Evaluators    : gpt-5-nano + Gemma-3N-E4B (no self-eval)")
print(f"Profiles      : E1 Strict | E2 Balanced | E3 Practical")
print(f"Total API calls: 50 × 3 × 2 = 300\n")

print(f"📂 Loading: {INPUT_CSV}")
df = pd.read_csv(INPUT_CSV, encoding="utf-8-sig")
df.columns = [c.strip() for c in df.columns]
df = df[df["rag_response"].notna()].reset_index(drop=True)
total = len(df)
print(f"✅ {total} RAG responses loaded\n")

# ============================================================
# STEP 2: Resume from checkpoint
# ============================================================
if os.path.exists(CHECKPOINT_CSV):
    existing = pd.read_csv(CHECKPOINT_CSV, encoding="utf-8-sig")
    done_ids = set(existing["question_id"].tolist())
    results  = existing.to_dict("records")
    print(f"♻️  Resuming — {len(done_ids)}/{total} already scored\n")
else:
    done_ids = set()
    results  = []
    print(f"🚀 Fresh start\n")

# ============================================================
# STEP 3: Score all questions
# ============================================================
start_time = time.time()

for i, row in df.iterrows():
    q_id     = row.get("question_id", f"Q{i+1}")
    tier     = row.get("tier", "")
    q_text   = str(row.get("question_text","")).strip()
    expected = str(row.get("expected_answer","")).strip()
    response = str(row.get("rag_response","")).strip()
    sources  = str(row.get("top_sources",""))

    if q_id in done_ids:
        print(f"[{i+1}/{total}] {q_id} ⏭ skipped")
        continue

    print(f"\n[{i+1}/{total}] {q_id} | Tier {tier}")

    entry = {
        "question_id":     q_id,
        "tier":            tier,
        "question_text":   q_text,
        "expected_answer": expected,
        "rag_response":    response,
        "top_sources":     sources,
    }

    profile_scores = {}

    for ev in EVALUATORS:
        eid        = ev["id"]
        ext_scores = []
        scorers    = []

        for scorer_cfg in EVALUATOR_MODELS:
            sname = scorer_cfg["name"]
            print(f"      {eid} ({ev['label']:35s}) → {sname} ...", end=" ", flush=True)
            scores, status = score_one(scorer_cfg, ev, q_text, expected, response)

            if status == "OK" and scores:
                ext_scores.append(scores)
                scorers.append(sname)
                print(f"✅ [{scores['factual']}+{scores['completeness']}+{scores['calibration']}={scores['total']}]")
            else:
                print(f"❌ {status}")
            time.sleep(DELAY)

        # Average the two scorers
        if len(ext_scores) == 2:
            avg_total = round(sum(s["total"] for s in ext_scores) / 2, 2)
            reasoning = f"avg({scorers[0]}:{ext_scores[0]['total']} + {scorers[1]}:{ext_scores[1]['total']})"
        elif len(ext_scores) == 1:
            avg_total = ext_scores[0]["total"]
            reasoning = f"only {scorers[0]} scored"
        else:
            avg_total = ""
            reasoning = "all scorers failed"

        entry[f"{eid}_total"]     = avg_total
        entry[f"{eid}_reasoning"] = reasoning
        if avg_total != "":
            profile_scores[eid] = avg_total
            print(f"      → {eid} score: {avg_total}/8")

    # Final avg across E1, E2, E3
    if profile_scores:
        final = round(sum(profile_scores.values()) / len(profile_scores), 2)
        entry["final_avg"] = final
        print(f"   ★ Final Avg: {final}/8")
    else:
        entry["final_avg"] = ""

    results.append(entry)
    pd.DataFrame(results).to_csv(CHECKPOINT_CSV, index=False)

    elapsed   = time.time() - start_time
    remaining = (elapsed / len(results)) * (total - len(results))
    print(f"   💾 Saved | ⏱ {elapsed/60:.1f}m | ~{remaining/60:.1f}m remaining")

results_df = pd.DataFrame(results)

# ============================================================
# CONSOLE SUMMARY
# ============================================================
print(f"\n{'='*60}")
print("  RAG EVALUATION COMPLETE")
print(f"{'='*60}")
overall = pd.to_numeric(results_df["final_avg"], errors="coerce").mean()
print(f"\n  gpt-oss-20b RAG  : {overall:.2f}/8")
print(f"  gpt-oss-20b Base : 5.45/8")
print(f"  Improvement      : {overall-5.45:+.2f}")
print()
BASELINE = {"1":4.37,"2":5.52,"3":5.42,"4":6.33,"5":5.60}
for tier in sorted(results_df["tier"].dropna().unique()):
    t_avg = pd.to_numeric(
        results_df[results_df["tier"]==tier]["final_avg"],
        errors="coerce").mean()
    base  = BASELINE.get(str(int(tier)), 5.45)
    print(f"  T{int(tier)}: {t_avg:.2f}/8  (baseline: {base}  Δ {t_avg-base:+.2f})")
print(f"\n{'='*60}\n")

# ============================================================
# WRITE EXCEL
# ============================================================
print(f"💾 Writing: {OUTPUT_EXCEL}")

wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "RAG Scores"

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
WHITE_BOLD= Font(name="Arial", bold=True, color="FFFFFF", size=11)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP      = Alignment(wrap_text=True, vertical="top")
TIER_CLR  = {"1":"EBF5FB","2":"E9F7EF","3":"FEF9E7","4":"FEF5E7","5":"FDEDEC"}
RAG_CLR   = {"hdr":"1A5276","resp":"D6EAF8","score":"AED6F1","final":"2E86C1"}
BASE_HDR  = "1F4E79"
def mf(h): return PatternFill("solid", fgColor=h)

# Header row
HEADERS = [
    ("question_id",    BASE_HDR),
    ("tier",           BASE_HDR),
    ("question_text",  BASE_HDR),
    ("expected_answer",BASE_HDR),
    ("rag_response",   RAG_CLR["hdr"]),
    ("top_sources",    RAG_CLR["hdr"]),
    ("E1 Strict /8",   RAG_CLR["hdr"]),
    ("E2 Balanced /8", RAG_CLR["hdr"]),
    ("E3 Practical /8",RAG_CLR["hdr"]),
    ("★ Final /8",     RAG_CLR["hdr"]),
]
for c,(label,color) in enumerate(HEADERS, start=1):
    cell=ws.cell(row=1, column=c, value=label)
    cell.font=HDR_FONT; cell.fill=mf(color); cell.alignment=CENTER
ws.row_dimensions[1].height = 30

DATA_CONFIG = [
    ("question_id",    None,            Font(name="Arial",bold=True,size=10), CENTER),
    ("tier",           None,            Font(name="Arial",bold=True,size=10), CENTER),
    ("question_text",  None,            BODY_FONT, WRAP),
    ("expected_answer",None,            BODY_FONT, WRAP),
    ("rag_response",   RAG_CLR["resp"], BODY_FONT, WRAP),
    ("top_sources",    RAG_CLR["resp"], BODY_FONT, WRAP),
    ("E1_total",       RAG_CLR["score"],BODY_FONT, CENTER),
    ("E2_total",       RAG_CLR["score"],BODY_FONT, CENTER),
    ("E3_total",       RAG_CLR["score"],BODY_FONT, CENTER),
    ("final_avg",      RAG_CLR["final"],WHITE_BOLD,CENTER),
]
for ri, row in results_df.iterrows():
    er    = ri + 2
    t_str = str(row.get("tier","")).strip()
    t_fill= mf(TIER_CLR.get(t_str,"FFFFFF"))
    for ci,(key,fill,font,align) in enumerate(DATA_CONFIG, start=1):
        val  = row.get(key,""); val = "" if pd.isna(val) else val
        cell = ws.cell(row=er, column=ci, value=val)
        cell.font=font; cell.fill=mf(fill) if fill else t_fill; cell.alignment=align

WIDTHS={1:16,2:6,3:42,4:38,5:55,6:40,7:13,8:14,9:14,10:12}
for c,w in WIDTHS.items():
    ws.column_dimensions[get_column_letter(c)].width=w
ws.freeze_panes="A2"

# Summary tab
ws2 = wb.create_sheet("Summary vs Baseline")
sum_hdrs=["Tier","Questions","E1 Strict","E2 Balanced","E3 Practical",
          "★ RAG Final","Baseline","Δ Improvement"]
for c,h in enumerate(sum_hdrs,start=1):
    cell=ws2.cell(row=1,column=c,value=h)
    cell.font=HDR_FONT; cell.fill=mf(BASE_HDR); cell.alignment=CENTER
ws2.row_dimensions[1].height=28

for ri,tier in enumerate(tiers := sorted(results_df["tier"].dropna().unique()),start=2):
    t_df = results_df[results_df["tier"]==tier]
    fill = mf(TIER_CLR.get(str(int(tier)),"FFFFFF"))
    e1   = pd.to_numeric(t_df["E1_total"],errors="coerce").mean()
    e2   = pd.to_numeric(t_df["E2_total"],errors="coerce").mean()
    e3   = pd.to_numeric(t_df["E3_total"],errors="coerce").mean()
    rag  = pd.to_numeric(t_df["final_avg"],errors="coerce").mean()
    base = BASELINE.get(str(int(tier)),5.45)
    delta= round(rag-base,2) if not pd.isna(rag) else ""
    vals = [f"T{int(tier)}",len(t_df),
            round(e1,2) if not pd.isna(e1) else "",
            round(e2,2) if not pd.isna(e2) else "",
            round(e3,2) if not pd.isna(e3) else "",
            round(rag,2) if not pd.isna(rag) else "",
            base, delta]
    for c,v in enumerate(vals,start=1):
        cell=ws2.cell(row=ri,column=c,value=v)
        cell.font=BODY_FONT; cell.fill=fill; cell.alignment=CENTER

ov_r   = len(tiers)+2
ov_rag = pd.to_numeric(results_df["final_avg"],errors="coerce").mean()
ov_vals= ["OVERALL",50,"","","",
          round(ov_rag,2),5.45,round(ov_rag-5.45,2)]
for c,v in enumerate(ov_vals,start=1):
    cell=ws2.cell(row=ov_r,column=c,value=v)
    cell.font=Font(name="Arial",bold=True,size=10)
    cell.fill=mf("D5D8DC"); cell.alignment=CENTER
for c in range(1,len(sum_hdrs)+1):
    ws2.column_dimensions[get_column_letter(c)].width=16
ws2.freeze_panes="A2"

wb.save(OUTPUT_EXCEL)
print(f"✅ Saved: {OUTPUT_EXCEL}")
print(f"✅ Checkpoint: {CHECKPOINT_CSV}")
print(f"\n🎉 Phase 2 Evaluation complete!")
